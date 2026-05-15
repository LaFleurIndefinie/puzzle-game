#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
报表需求确认单 Excel 生成器（模板格式）

从 SQL 和 entities.json 自动生成确认单 Excel，包含：
- 业务对象信息（一、业务对象确认）
- 字段映射信息（二、输出字段确认）
- JOIN关系说明（三、JOIN关系说明）
- 查询条件说明（四、查询条件说明）

Usage:
    from gen_confirm_excel import generate_confirm_excel
    generate_confirm_excel(sql_content, entities_json_path, output_path, report_name)

    python gen_confirm_excel.py --sql-file report.sql --entities entities.json --output confirm.xlsx --report "报表名"
"""

from __future__ import annotations

import json
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple


# ============================================================
# SQL 解析模块
# ============================================================

def _extract_joins(search_range: str, tables: List[Dict], joins: List[Dict], check_dup_type: str = 'main'):
    """从指定范围提取 JOIN 表和 ON 条件"""
    join_pattern = r'(LEFT|RIGHT|INNER)?\s*JOIN\s+([^\s()]+)\.([^\s(,]+)\s+(?:AS\s+)?(\w+)'
    # 已知业务表别名白名单（过滤 hist* 等子查询辅助表）
    known_aliases = {'hdr', 'ast', 'dtl', 'cust', 'rev_hdr', 'body'}
    for match in re.finditer(join_pattern, search_range, re.IGNORECASE):
        join_type = match.group(1) or 'LEFT JOIN'
        schema = match.group(2).strip()
        table = match.group(3).strip()
        alias = match.group(4).strip()
        # 跳过 hist* 子查询辅助表的 JOIN
        if alias.startswith('hist'):
            continue
        key = f"{schema}.{table}"
        # 检查是否已存在（避免重复添加）
        if not any(t.get('alias') == alias.strip() for t in tables):
            tables.append({'schema': schema, 'table': table, 'alias': alias.strip(), 'type': 'join'})
        # 提取 ON 条件
        on_start = match.end()
        on_window = search_range[on_start:on_start + 500]
        on_pattern = rf'\bON\s+(\w+)\.(\w+)\s*=\s*(\w+)\.(\w+)'
        on_match = re.search(on_pattern, on_window, re.IGNORECASE)
        if on_match:
            left_alias = on_match.group(1).strip()
            right_alias = on_match.group(3).strip()
            # 跳过涉及 hist* 的 JOIN
            if left_alias.startswith('hist') or right_alias.startswith('hist'):
                continue
            # 检查是否已存在相同的 JOIN
            exists = any(
                j.get('left_alias') == left_alias and
                j.get('left_field') == on_match.group(2).strip() and
                j.get('right_alias') == right_alias and
                j.get('right_field') == on_match.group(4).strip()
                for j in joins
            )
            if not exists:
                joins.append({
                    'type': join_type.strip(),
                    'left_alias': left_alias,
                    'left_field': on_match.group(2).strip(),
                    'right_alias': right_alias,
                    'right_field': on_match.group(4).strip(),
                    'note': ''
                })


def parse_sql_content(sql_content: str) -> Tuple[List[Dict], List[Dict], List[Dict], List[str]]:
    """从SQL中解析表、字段、JOIN关系和WHERE条件

    Returns:
        tables: [{'schema': str, 'table': str, 'alias': str, 'type': str}]
        joins: [{'type': str, 'left_alias': str, 'left_field': str, 'right_alias': str, 'right_field': str, 'note': str}]
        fields: [{'alias': str, 'field': str, 'display_name': str, 'table_name': str, 'db_path': str}]
        conditions: [str]
    """
    tables = []
    joins = []
    fields = []
    conditions = []

    # 提取主 FROM 和 WHERE 之间的范围（主查询 JOIN 区域）
    def get_main_query_range(sql: str):
        """返回 (from_pos, where_pos) 主表 JOIN 范围"""
        # 主 FROM: FROM\n    yonbip_fi_ctmfc.stwb_settleapply hdr
        from_pos = sql.find('FROM\n    yonbip_fi_ctmfc.stwb_settleapply')
        if from_pos < 0:
            return -1, -1
        where_pos = sql.find('WHERE', from_pos)
        return from_pos, where_pos

    main_from_pos, main_where_pos = get_main_query_range(sql_content)

    # 首先提取主 FROM（明确的主表）
    if main_from_pos >= 0:
        # 匹配主 FROM: schema.tablename AS alias 或 schema.tablename alias
        main_from_pattern = r'FROM\n    ([^\s()]+)\.([^\s(),]+)\s+(?:AS\s+)?(\w+)'
        main_match = re.search(main_from_pattern, sql_content[main_from_pos:main_from_pos + 200], re.IGNORECASE)
        if main_match:
            schema, table, alias = main_match.groups()
            tables.append({'schema': schema.strip(), 'table': table.strip(), 'alias': alias.strip(), 'type': 'main'})

    # 提取 JOIN 表和 ON 条件（主查询范围 + 子查询范围）
    # 主查询 JOIN 范围：FROM 和 WHERE 之间
    if main_from_pos >= 0 and main_where_pos > main_from_pos:
        join_range = sql_content[main_from_pos:main_where_pos]
        _extract_joins(join_range, tables, joins, check_dup_type='main')

    # 子查询 JOIN 范围：整个 SQL（用于捕获 SELECT 子句中的 JOIN）
    _extract_joins(sql_content, tables, joins, check_dup_type='subquery')

    # 提取子查询中的表（FROM 子查询，支持有无 AS 两种形式）
    subquery_pattern = r'FROM\s+([^\s()]+)\.([^\s(,]+)\s+(?:AS\s+)?(\w+)'
    for match in re.finditer(subquery_pattern, sql_content, re.IGNORECASE):
        schema, table, alias = match.groups()
        if not any(t.get('alias') == alias.strip() for t in tables):
            tables.append({'schema': schema.strip(), 'table': table.strip(), 'alias': alias.strip(), 'type': 'subquery'})

    # 提取字段（从 SELECT ... AS "..."）
    # 支持三种模式：
    # 1. 普通字段: alias.field AS "中文名"
    # 2. 函数字段: SUM(alias.field) AS "中文名" / MAX(alias.field) AS "中文名"
    # 3. 子查询字段: (SELECT ...) AS "中文名"

    # 模式1: 普通字段 alias.field AS "中文名"
    simple_pattern = r'(\w+)\.(\w+)\s+AS\s+"([^"]+)"'
    for match in re.finditer(simple_pattern, sql_content):
        alias, field, display_name = match.groups()
        table_info = next((t for t in tables if t.get('alias') == alias.strip()), None)
        table_name = table_info['table'] if table_info else ''
        fields.append({
            'alias': alias.strip(),
            'field': field.strip(),
            'display_name': display_name.strip(),
            'table_name': table_name,
            'db_path': f"{table_name}#{field.strip()}" if table_name else field.strip()
        })

    # 模式2: 函数字段 SUM(alias.field) AS "中文名" 等
    func_pattern = r'(?:SUM|MAX|MIN|COUNT|AVG)\s*\(\s*(\w+)\.(\w+)\s*\)\s+AS\s+"([^"]+)"'
    for match in re.finditer(func_pattern, sql_content, re.IGNORECASE):
        alias, field, display_name = match.groups()
        table_info = next((t for t in tables if t.get('alias') == alias.strip()), None)
        table_name = table_info['table'] if table_info else ''
        # 避免重复添加（已在模式1中添加的）
        if not any(f.get('display_name') == display_name.strip() for f in fields):
            fields.append({
                'alias': alias.strip(),
                'field': field.strip(),
                'display_name': display_name.strip(),
                'table_name': table_name,
                'db_path': f"{table_name}#{field.strip()}" if table_name else field.strip()
            })

    # 模式3: 子查询字段 (SELECT ...) AS "中文名"
    # 找所有 ) AS "..." 模式，但需要验证开括号后是 SELECT（而非函数名）
    subquery_pattern = r'\)\s+AS\s+"([^"]+)"'
    for match in re.finditer(subquery_pattern, sql_content):
        display_name = match.group(1).strip()
        end_pos = match.start()

        # 向前找到匹配的 (
        depth = 1
        pos = end_pos - 1
        while pos >= 0 and depth > 0:
            if sql_content[pos] == ')':
                depth += 1
            elif sql_content[pos] == '(':
                depth -= 1
            pos -= 1
        open_pos = pos + 1

        # 检查 ( 之后是否紧跟 SELECT（忽略空白）
        inner_start = open_pos + 1
        while inner_start < end_pos and sql_content[inner_start] in ' \t\n\r':
            inner_start += 1

        if sql_content[inner_start:inner_start+6].upper() == 'SELECT' and not any(f.get('display_name') == display_name for f in fields):
            # 这是子查询字段
            # 从 lookback 中提取表信息
            lookback = sql_content[open_pos:match.start()]
            # 尝试找 table.field 或 table.field AS 模式
            table_field_pat = r'(\w+)\.(\w+)'
            tf_matches = re.findall(table_field_pat, lookback, re.IGNORECASE)
            alias = field = ''
            table_name = ''
            if tf_matches:
                alias, field = tf_matches[-1]
                # 解析 alias 为实际表名
                alias = alias.strip()
                field = field.strip()
                table_info = next((t for t in tables if t.get('alias') == alias), None)
                table_name = table_info['table'] if table_info else alias
            fields.append({
                'alias': alias,
                'field': field if field else 'subquery',
                'display_name': display_name,
                'table_name': table_name,
                'db_path': f"子查询: {display_name[:20]}" if not field else ''
            })

    # 模式4: CASE WHEN ... END AS "中文名"
    case_when_pattern = r'END\s+AS\s+"([^"]+)"'
    for match in re.finditer(case_when_pattern, sql_content):
        display_name = match.group(1).strip()
        if not any(f.get('display_name') == display_name for f in fields):
            fields.append({
                'alias': '',
                'field': 'CASE WHEN',
                'display_name': display_name,
                'table_name': '',
                'db_path': f"CASE WHEN: {display_name[:20]}"
            })

    # 模式5: 常量 AS "中文名" (如 0 AS "剩余未冲销金额")
    constant_pattern = r'\b0\s+AS\s+"([^"]+)"'
    for match in re.finditer(constant_pattern, sql_content):
        display_name = match.group(1).strip()
        if not any(f.get('display_name') == display_name for f in fields):
            fields.append({
                'alias': '',
                'field': '0',
                'display_name': display_name,
                'table_name': '',
                'db_path': '0 (常量)'
            })

    # 提取主查询的 WHERE 条件
    def extract_main_where_conditions(sql: str) -> List[str]:
        """提取主查询的 WHERE 条件（主 FROM 之后的 WHERE，不是子查询中的 WHERE）"""
        from_pos = sql.find('FROM\n    yonbip')
        if from_pos < 0:
            return []
        where_pos = sql.find('WHERE', from_pos)
        if where_pos < 0:
            return []
        group_pos = sql.find('GROUP BY', where_pos)
        if group_pos < 0:
            return []
        where_text = sql[where_pos + 5:group_pos]
        depth = 0
        parts = []
        current = []
        i = 0
        while i < len(where_text):
            ch = where_text[i]
            if ch == '(':
                depth += 1
            elif ch == ')':
                depth -= 1
            elif depth == 0 and i + 2 < len(where_text) and where_text[i:i+3].upper() == 'AND':
                cond = ''.join(current).strip()
                cond = re.sub(r'^\s*(AND|OR)\s+', '', cond, flags=re.IGNORECASE).strip()
                if cond:
                    parts.append(cond)
                current = []
                i += 3
                continue
            current.append(ch)
            i += 1
        cond = ''.join(current).strip()
        cond = re.sub(r'^\s*(AND|OR)\s+', '', cond, flags=re.IGNORECASE).strip()
        if cond:
            parts.append(cond)
        result = []
        for p in parts:
            p = p.strip()
            p = re.sub(r'--.*$', '', p, flags=re.MULTILINE).strip()
            p = re.sub(r'/\*.*?\*/', '', p, flags=re.DOTALL).strip()
            if p and p not in ('WHERE', ''):
                result.append(p)
        return result

    conditions = extract_main_where_conditions(sql_content)

    # GROUP BY 字段（仅取主 GROUP BY）
    def extract_main_group_by(sql: str) -> str:
        from_pos = sql.find('FROM\n    yonbip')
        if from_pos < 0:
            return ''
        group_pos = sql.find('GROUP BY', from_pos)
        if group_pos < 0:
            return ''
        # GROUP BY 跨多行，提取从 "GROUP BY" 到 SQL 末尾的内容
        # 清理多余空白，换行替换为逗号分隔
        content = sql[group_pos + 8:].strip()
        # 移除前后空白
        content = re.sub(r'\s+', ' ', content).strip()
        return content

    main_group = extract_main_group_by(sql_content)
    if main_group:
        conditions.append(f"GROUP BY: {main_group}")

    return tables, joins, fields, conditions


def get_entity_info(entities_json_path: str, table_name: str) -> Dict[str, str]:
    """从 entities.json 中获取表对应的业务对象信息"""
    try:
        with open(entities_json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        for entity in data.get('entities', []):
            if entity.get('tableName') == table_name:
                return {
                    'billName': entity.get('billName', ''),
                    'schema': entity.get('schema', ''),
                    'uri': entity.get('uri', ''),
                    'description': ''
                }
    except Exception:
        pass
    return {'billName': '', 'schema': '', 'uri': '', 'description': ''}


def get_field_display_name(entities_json_path: str, table_name: str, field_name: str) -> str:
    """从 entities.json 中获取字段的 displayName"""
    try:
        with open(entities_json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        for entity in data.get('entities', []):
            if entity.get('tableName') == table_name:
                for attr in entity.get('attributes', []):
                    if attr.get('dbColumnName') == field_name:
                        return attr.get('displayName', '')
    except Exception:
        pass
    return ''


def determine_field_type(alias: str, field: str, raw_expr: str = '') -> Tuple[str, str]:
    """判断字段类型和补充说明

    Returns:
        (field_type, note)
    """
    if field == '0' and raw_expr.startswith('0'):
        return '计算字段', '预留为0'
    if 'CASE' in field.upper() or 'IFNULL' in field.upper():
        if 'EXISTS' in field.upper():
            return '条件计算', 'CASE WHEN EXISTS'
        return '计算字段', ''
    if raw_expr.startswith('='):
        return '计算字段', raw_expr[:30]
    if '(' in field and ')' in field:
        return '子查询', '嵌套子查询'
    return '普通字段', ''


# ============================================================
# Excel 生成模块（模板格式）
# ============================================================

def generate_confirm_excel(
    sql_content: str,
    entities_json_path: str,
    output_path: str,
    report_name: str = ""
) -> bool:
    """生成报表需求确认单 Excel（模板格式）

    Args:
        sql_content: SQL 内容
        entities_json_path: entities.json 文件路径
        output_path: 输出 Excel 文件路径
        report_name: 报表名称

    Returns:
        bool: 是否成功生成
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("错误: 需要安装 openpyxl 库", file=sys.stderr)
        return False

    # 解析 SQL
    tables, joins, fields, conditions = parse_sql_content(sql_content)

    # 去重表（保持顺序，过滤 hist* 子查询辅助表）
    seen = set()
    unique_tables = []
    for t in tables:
        # 跳过 hist* 子查询辅助表
        if t.get('alias', '').startswith('hist'):
            continue
        key = f"{t['schema']}.{t['table']}"
        if key not in seen:
            seen.add(key)
            unique_tables.append(t)

    # 获取每个表的 entity 信息
    for t in unique_tables:
        ei = get_entity_info(entities_json_path, t['table'])
        t['billName'] = ei.get('billName', '')
        t['uri'] = ei.get('uri', '')
        t['description'] = ei.get('description', '')

    # 建立 table_name -> (billName, uri) 的映射
    table_to_info = {}
    for t in unique_tables:
        table_to_info[t['table']] = {'billName': t.get('billName', ''), 'uri': t.get('uri', '')}

    # 获取每个字段的 displayName 和正确的 db_path (uri#字段名 格式)
    for f in fields:
        if f['table_name']:
            f['meta_name'] = get_field_display_name(entities_json_path, f['table_name'], f['field'])
            # Col 4 = 业务对象名#字段名 格式
            bill_name = table_to_info.get(f['table_name'], {}).get('billName', f['table_name'])
            f['db_path'] = f"{bill_name}#{f['field']}"
            # Col 5 = 字段名(uri#字段名) 格式
            uri = table_to_info.get(f['table_name'], {}).get('uri', '')
            uri_path = f"{uri}#{f['field']}" if uri else f['db_path']
            f['uri_path'] = f"{f['display_name']}({uri_path})"
            f['source_table'] = f['table_name']
            f['source_alias'] = f['alias']
        else:
            f['meta_name'] = ''
            f['source_table'] = '-'
            f['source_alias'] = '-'
            f['db_path'] = f['display_name']
            f['uri_path'] = ''

    # ---- 样式定义 ----
    HEADER_FILL = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    HEADER_FONT = Font(name='Arial', bold=True, size=11)
    DATA_FONT = Font(name='Arial', size=10)
    TITLE_FONT = Font(name='Arial', bold=True, size=12)
    DESC_FONT = Font(name='Arial', size=10)
    PENDING_FONT = Font(name='Arial', size=10, color='FF0000')
    CORRECT_FONT = Font(name='Arial', size=10, color='0070C0')
    BORDER = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LEFT_WRAP = Alignment(horizontal='left', vertical='center', wrap_text=True)
    LEFT = Alignment(horizontal='left', vertical='center')

    def hdr_cell(ws, row, col, value):
        c = ws.cell(row=row, column=col, value=value)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
        return c

    def data_cell(ws, row, col, value, font=None, alignment=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font = font or DATA_FONT
        c.alignment = alignment or LEFT_WRAP
        c.border = BORDER
        return c

    # ---- 创建工作簿 ----
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '确认单'

    r = 1

    # === Row 1: Title ===
    ws.merge_cells(f'A{r}:H{r}')
    title_val = f"报表9  {report_name}" if report_name else "报表9"
    c = ws.cell(row=r, column=1, value=title_val)
    c.font = TITLE_FONT
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[r].height = 18
    r += 1

    # === Row 2: Description ===
    ws.merge_cells(f'A{r}:H{r}')
    c = ws.cell(row=r, column=1, value='说明：表格用于筛选、去重及汇总；为汇总数据提供维度支撑；为交叉筛选提供维度配置')
    c.font = DESC_FONT
    c.alignment = LEFT
    ws.row_dimensions[r].height = 16
    r += 1

    # === Section 1: 业务对象确认 ===
    ws.merge_cells(f'A{r}:H{r}')
    hdr_cell(ws, r, 1, '一、业务对象确认')
    ws.row_dimensions[r].height = 18
    r += 1

    # 业务对象表头
    biz_headers = ['序号', '原业务对象名称', 'BIP业务对象实体名称', '实体说明', 'database schema', 'database tableName', '单元化URI']
    for col, h in enumerate(biz_headers, 1):
        hdr_cell(ws, r, col, h)
    ws.row_dimensions[r].height = 18
    r += 1

    # 业务对象数据行
    for idx, t in enumerate(unique_tables, 1):
        desc = t.get('description', '') or t.get('billName', '')
        data_cell(ws, r, 1, idx, alignment=CENTER)
        data_cell(ws, r, 2, t.get('billName', ''), alignment=LEFT)
        data_cell(ws, r, 3, t.get('billName', ''), alignment=LEFT)
        data_cell(ws, r, 4, desc, alignment=LEFT_WRAP)
        data_cell(ws, r, 5, t['schema'], alignment=CENTER)
        data_cell(ws, r, 6, t['table'], alignment=CENTER)
        data_cell(ws, r, 7, t.get('uri', ''), alignment=LEFT_WRAP)
        ws.row_dimensions[r].height = 28
        r += 1

    r += 1  # 空行

    # === Section 2: 输出字段确认 ===
    # 模板格式：Row 11 = section header (A merged) + column headers (C, D, ...)
    # 先写 section title（合并A列），再写 column headers
    ws.merge_cells(f'A{r}:H{r}')
    hdr_cell(ws, r, 1, '二、输出字段确认')
    ws.row_dimensions[r].height = 18
    r += 1

    field_headers = ['序号', '', '输出字段', '数据库字段(表名#字段名格式)', '数据库实体字段含义', '来源表', '来源标识', '补充说明']
    for col, h in enumerate(field_headers, 1):
        hdr_cell(ws, r, col, h)
    ws.row_dimensions[r].height = 18
    r += 1

    # 字段数据行
    for idx, f in enumerate(fields, 1):
        ftype, fnote = determine_field_type(f['alias'], f['field'])
        note = fnote
        font = DATA_FONT
        if '待' in note or '预留' in note:
            font = PENDING_FONT

        data_cell(ws, r, 1, idx, alignment=CENTER)
        data_cell(ws, r, 2, '', alignment=CENTER)
        data_cell(ws, r, 3, f['display_name'], alignment=LEFT_WRAP)
        data_cell(ws, r, 4, f['db_path'], alignment=LEFT_WRAP)
        data_cell(ws, r, 5, f.get('uri_path', ''), alignment=LEFT_WRAP)
        data_cell(ws, r, 6, f['source_table'], alignment=CENTER)
        data_cell(ws, r, 7, f['source_alias'], alignment=CENTER)
        data_cell(ws, r, 8, note, font=font, alignment=LEFT_WRAP)
        ws.row_dimensions[r].height = 28
        r += 1

    r += 1  # 空行

    # === Section 3: JOIN关系说明 ===
    section3_hdr_row = r
    ws.merge_cells(f'A{r}:H{r}')
    hdr_cell(ws, r, 1, '三、JOIN关系说明')
    ws.row_dimensions[r].height = 18
    r += 1

    join_headers = ['序号', '', 'JOIN类型', '左表', '左表字段', '右表', '右表字段', '补充说明']
    for col, h in enumerate(join_headers, 1):
        hdr_cell(ws, r, col, h)
    ws.row_dimensions[r].height = 18
    ws.merge_cells(f'A{section3_hdr_row}:A{r}')  # A列合并：section标题行+列头行
    r += 1

    for idx, j in enumerate(joins, 1):
        # 查找左右表的表名
        left_table = next((t['table'] for t in unique_tables if t['alias'] == j['left_alias']), j['left_alias'])
        right_table = next((t['table'] for t in unique_tables if t['alias'] == j['right_alias']), j['right_alias'])

        data_cell(ws, r, 1, idx, alignment=CENTER)
        data_cell(ws, r, 2, '', alignment=CENTER)
        data_cell(ws, r, 3, j['type'], alignment=CENTER)
        data_cell(ws, r, 4, f"{left_table}({j['left_alias']})", alignment=LEFT_WRAP)
        data_cell(ws, r, 5, j['left_field'], alignment=CENTER)
        data_cell(ws, r, 6, f"{right_table}({j['right_alias']})", alignment=LEFT_WRAP)
        data_cell(ws, r, 7, j['right_field'], alignment=CENTER)
        data_cell(ws, r, 8, j.get('note', ''), alignment=LEFT_WRAP)
        ws.row_dimensions[r].height = 28
        r += 1

    r += 1  # 空行

    # === Section 4: 查询条件说明 ===
    ws.merge_cells(f'A{r}:H{r}')
    hdr_cell(ws, r, 1, '四、查询条件说明')
    ws.row_dimensions[r].height = 18
    r += 1

    for cond in conditions:
        ws.merge_cells(f'A{r}:H{r}')
        c = ws.cell(row=r, column=1, value=cond)
        c.font = DATA_FONT
        c.alignment = LEFT
        c.border = BORDER
        ws.row_dimensions[r].height = 22
        r += 1

    # === 列宽设置 ===
    col_widths = {'A': 8, 'B': 23, 'C': 27, 'D': 29, 'E': 47, 'F': 18, 'G': 38, 'H': 45}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # === 保存 ===
    wb.save(output_path)
    return True


# ============================================================
# 主入口（支持命令行调用）
# ============================================================

def main():
    import argparse

    parser = argparse.ArgumentParser(
        description="报表需求确认单 Excel 生成器（模板格式）",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
    python gen_confirm_excel.py --sql-file report.sql --entities entities.json --output confirm.xlsx --report "热费收款冲暂估表"
        """
    )
    parser.add_argument('--sql-file', required=True, help="SQL 文件路径")
    parser.add_argument('--entities', required=True, help="entities.json 文件路径")
    parser.add_argument('--output', required=True, help="输出 Excel 文件路径")
    parser.add_argument('--report', default='', help="报表名称")

    args = parser.parse_args()

    sql_path = Path(args.sql_file)
    if not sql_path.exists():
        print(f"错误: SQL 文件不存在: {args.sql_file}", file=sys.stderr)
        return 1

    sql_content = sql_path.read_text(encoding='utf-8')

    print(f"正在生成确认单: {args.output}")
    if generate_confirm_excel(sql_content, args.entities, args.output, args.report):
        print(f"成功: {args.output}")
        return 0
    else:
        print("生成失败", file=sys.stderr)
        return 1


if __name__ == "__main__":
    sys.exit(main())
