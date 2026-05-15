#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
字段映射表转 Excel 生成器

根据阶段二输出的字段映射表（Markdown格式）反向生成 Excel 文件，供用户确认字段映射关系。

Usage:
    python field_mapping_to_excel.py output/字段映射表_核销进度.md
    python field_mapping_to_excel.py output/字段映射表_核销进度.md -o output/字段映射表_核销进度.xlsx
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple


def parse_markdown_table(content: str) -> List[List[str]]:
    """解析 Markdown 表格，返回二维列表"""
    lines = content.strip().split('\n')
    rows = []

    for line in lines:
        line = line.strip()
        if not line or line.startswith('#') or line.startswith('>'):
            continue
        # 匹配 Markdown 表格行（以 | 开头和结尾）
        if line.startswith('|') and line.endswith('|'):
            # 跳过表头分隔行（如 |------|------|）
            if re.match(r'^\|[\s\-:|]+\|$', line):
                continue
            # 提取单元格内容
            cells = [cell.strip() for cell in line.split('|')[1:-1]]
            rows.append(cells)

    return rows


def extract_field_mapping_from_markdown(md_content: str) -> Dict[str, Any]:
    """从 Markdown 内容中提取字段映射信息"""
    result = {
        'report_name': '',
        'business_objects': [],
        'main_table': '',
        'schema': '',
        'fields': []
    }

    # 提取报表名称
    match = re.search(r'#\s*(.+)', md_content)
    if match:
        result['report_name'] = match.group(1).strip()

    # 提取业务对象信息表
    lines = md_content.split('\n')
    in_field_section = False
    header = []

    for i, line in enumerate(lines):
        line = line.strip()

        # 检测字段映射表开始
        if '## 三、字段映射表' in line or '## 三、字段映射' in line:
            in_field_section = True
            continue

        # 检测字段映射表结束
        if in_field_section and line.startswith('## '):
            break

        # 检测业务对象信息表
        if '## 二、业务对象信息' in line:
            # 解析接下来的表格
            table_rows = []
            for j in range(i + 1, min(i + 15, len(lines))):
                l = lines[j].strip()
                if l.startswith('|') and l.endswith('|') and '|--' not in l:
                    cells = [c.strip() for c in l.split('|')[1:-1]]
                    table_rows.append(cells)
                if not l.startswith('|'):
                    break

            # 提取业务对象信息
            if table_rows and len(table_rows) > 1:
                for row in table_rows[1:]:  # 跳过表头
                    if len(row) >= 5:
                        obj = {
                            'description': row[1] if len(row) > 1 else '',
                            'bip_name': row[2] if len(row) > 2 else '',
                            'type': row[3] if len(row) > 3 else '',
                            'schema': row[4] if len(row) > 4 else '',
                            'table': row[5] if len(row) > 5 else '',
                        }
                        result['business_objects'].append(obj)
                        # 记录主表信息
                        if '主实体' in obj['type'] or '主表' in obj['type']:
                            result['main_table'] = obj['table']
                            result['schema'] = obj['schema']

        # 解析字段映射表
        if in_field_section:
            if not header and '|' in line:
                header = [c.strip() for c in line.split('|')[1:-1]]
                continue

            if header and '|' in line and '|--' not in line:
                cells = [c.strip() for c in line.split('|')[1:-1]]
                if len(cells) >= 4:
                    field = {
                        'seq': cells[0] if len(cells) > 0 else '',
                        'report_field': cells[1] if len(cells) > 1 else '',
                        'field_source': cells[2] if len(cells) > 2 else '',
                        'table_name': cells[3] if len(cells) > 3 else '',
                        'schema': cells[4] if len(cells) > 4 else '',
                        'db_column': cells[5] if len(cells) > 5 else '',
                        'field_type': cells[6] if len(cells) > 6 else '',
                        'join_desc': cells[7] if len(cells) > 7 else '',
                    }
                    result['fields'].append(field)

    return result


def create_excel(data: Dict[str, Any], output_path: str) -> bool:
    """创建 Excel 文件"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("错误: 需要安装 openpyxl 库", file=sys.stderr)
        print("请运行: pip install openpyxl", file=sys.stderr)
        return False

    wb = openpyxl.Workbook()

    # ========== Sheet 1: 字段映射表 ==========
    ws1 = wb.active
    ws1.title = "字段映射"

    # 样式定义
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 报表信息区域
    ws1['A1'] = "报表名称"
    ws1['B1'] = data['report_name']
    ws1['A1'].font = Font(bold=True)
    ws1['A2'] = "主表"
    ws1['B2'] = f"{data['schema']}.{data['main_table']}" if data['main_table'] else ""

    # 字段映射表头
    headers = ["序号", "报表字段", "字段来源", "表名", "Schema", "数据库字段", "字段类型", "关联说明"]
    header_row = 4
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 填充字段数据
    for row_idx, field in enumerate(data['fields'], header_row + 1):
        values = [
            field.get('seq', ''),
            field.get('report_field', ''),
            field.get('field_source', ''),
            field.get('table_name', ''),
            field.get('schema', ''),
            field.get('db_column', ''),
            field.get('field_type', ''),
            field.get('join_desc', ''),
        ]
        for col, value in enumerate(values, 1):
            cell = ws1.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    # 设置列宽
    column_widths = [8, 15, 10, 25, 20, 20, 15, 30]
    for i, width in enumerate(column_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = width

    # ========== Sheet 2: 业务对象信息 ==========
    ws2 = wb.create_sheet(title="业务对象")

    headers2 = ["原业务描述", "BIP实际名称", "实体说明", "Schema", "表名", "URI"]
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, obj in enumerate(data['business_objects'], 2):
        values = [
            obj.get('description', ''),
            obj.get('bip_name', ''),
            obj.get('type', ''),
            obj.get('schema', ''),
            obj.get('table', ''),
        ]
        for col, value in enumerate(values, 1):
            cell = ws2.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border

    # 设置业务对象表列宽
    widths2 = [15, 15, 12, 25, 30, 50]
    for i, width in enumerate(widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = width

    # ========== Sheet 3: SQL 参考 ==========
    ws3 = wb.create_sheet(title="SQL参考")

    # 生成 JOIN 语句参考
    joins = []
    for obj in data['business_objects']:
        if '主实体' in obj['type'] or '主表' in obj['type']:
            continue  # 主表已经在 FROM 中
        schema = obj.get('schema', '')
        table = obj.get('table', '')
        obj_type = obj.get('type', '')

        if '参照' in obj_type:
            joins.append(f"LEFT JOIN {schema}.{table} ON ...")
        elif '子' in obj_type:
            joins.append(f"LEFT JOIN {schema}.{table} ON ...")
        elif '特征' in obj_type:
            joins.append(f"LEFT JOIN {schema}.{table} ON ...")

    ws3['A1'] = "JOIN 语句参考"
    ws3['A1'].font = Font(bold=True, size=12)
    for row_idx, join_sql in enumerate(joins, 3):
        ws3.cell(row=row_idx, column=1, value=join_sql)

    ws3.column_dimensions['A'].width = 60

    # 保存文件
    wb.save(output_path)
    return True


def main():
    parser = argparse.ArgumentParser(
        description="字段映射表转 Excel 生成器",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
    python field_mapping_to_excel.py output/字段映射表.md
    python field_mapping_to_excel.py output/字段映射表.md -o output/确认单.xlsx
        """
    )
    parser.add_argument(
        'input',
        help="字段映射表 Markdown 文件路径"
    )
    parser.add_argument(
        '-o', '--output',
        help="输出 Excel 文件路径（默认: 字段映射表.xlsx）"
    )

    args = parser.parse_args()

    # 读取输入文件
    input_path = Path(args.input)
    if not input_path.exists():
        print(f"错误: 文件不存在: {input_path}", file=sys.stderr)
        sys.exit(1)

    with open(input_path, 'r', encoding='utf-8') as f:
        md_content = f.read()

    # 解析 Markdown
    data = extract_field_mapping_from_markdown(md_content)

    if not data['fields']:
        print("警告: 未找到字段映射数据", file=sys.stderr)

    # 确定输出路径
    if args.output:
        output_path = args.output
    else:
        output_path = str(input_path.parent / f"{input_path.stem}.xlsx")

    # 生成 Excel
    print(f"正在生成 Excel: {output_path}")
    if create_excel(data, output_path):
        print(f"成功: {output_path}")
        print(f"  - 报表名称: {data['report_name']}")
        print(f"  - 字段数量: {len(data['fields'])}")
        print(f"  - 业务对象: {len(data['business_objects'])}")
    else:
        sys.exit(1)


if __name__ == "__main__":
    main()
