#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
字段预检验脚本 v2.0（合并版）

合并 phase1(field_verify.py) 和 phase2(validate_fields.py) 的能力：
1. 字段发现：根据显示名/Excel列头查找元数据中的 dbColumnName（支持参照路径解析）
2. 字段校验：验证映射表中的 tableName.dbColumnName 是否在元数据中真实存在
3. WHERE 校验：验证 WHERE 条件字段是否出现在 SELECT 中
4. 模糊匹配：字段不存在时，Levenshtein 距离推荐相似字段

Usage:
    # 方式1：从 Excel 提取字段并验证
    python field_verify.py --entities output/entities.json --excel 报表模板.xlsx

    # 方式2：验证字段列表
    python field_verify.py --entities output/entities.json --fields 收款日期 收款金额 客户名称

    # 方式3：验证字段映射表（md）
    python field_verify.py --entities output/entities.json --mapping output/字段映射表.md

    # 方式4：校验 WHERE 字段是否在 SELECT 中
    python field_verify.py --mapping output/字段映射表.md --check-where

    # 方式5：完整校验（映射表 + WHERE + 字段验证）
    python field_verify.py --entities output/entities.json --mapping output/字段映射表.md --check-where
"""

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple


# =============================================================================
# 第一部分：元数据加载与索引构建
# =============================================================================

def load_entities(entities_path: str) -> List[Dict[str, Any]]:
    """加载 entities.json"""
    with open(entities_path, 'r', encoding='utf-8') as f:
        return json.load(f).get('entities', [])


def build_column_index(entities: List[Dict[str, Any]]) -> Dict[str, Set[str]]:
    """
    构建字段索引：schema.tableName -> set(dbColumnName)

    用于快速校验已知 tableName.dbColumnName 是否在元数据中存在。
    """
    index: Dict[str, Set[str]] = {}
    for entity in entities:
        schema = entity.get('schema', '')
        table = entity.get('tableName', '')
        if not table:
            continue
        key = f"{schema}.{table}" if schema else table

        columns: Set[str] = set()
        for attr in entity.get('attributes', []):
            col = attr.get('dbColumnName', '')
            if col:
                columns.add(col)

        if key not in index:
            index[key] = columns
        else:
            index[key].update(columns)

    return index


def build_displayname_index(entities: List[Dict[str, Any]]) -> Dict[str, List[Dict]]:
    """
    构建 displayName -> [(entity, attr), ...] 的反向索引

    用于快速根据显示名查找字段。
    """
    index: Dict[str, List[Dict]] = {}
    for entity in entities:
        table_name = entity.get('tableName', '')
        uri = entity.get('uri', '')
        bill_name = entity.get('billName', '')
        for attr in entity.get('attributes', []):
            dn = attr.get('displayName', '')
            if not dn:
                continue
            if dn not in index:
                index[dn] = []
            index[dn].append({
                'dbColumnName': attr.get('dbColumnName'),
                'tableName': table_name,
                'entityUri': uri,
                'billName': bill_name,
                'type': attr.get('type'),
                'uri': attr.get('uri'),
            })
    return index


# =============================================================================
# 第二部分：字段查找（发现能力）
# =============================================================================

def find_entity_by_table_name(entities: List[Dict], table_name: str) -> Optional[Dict]:
    """根据表名查找实体"""
    for entity in entities:
        if entity.get('tableName') == table_name:
            return entity
    return None


def find_entity_by_uri(entities: List[Dict], uri: str) -> Optional[Dict]:
    """根据 URI 查找实体"""
    for entity in entities:
        if entity.get('uri') == uri:
            return entity
    return None


def _guess_table_from_uri(uri: str) -> Optional[str]:
    """从 URI 猜测表名（兜底逻辑）"""
    parts = uri.split('.')
    return parts[-1] if parts else None


def _check_reference_field(
    entity: Dict,
    display_name: str,
    entities: List[Dict]
) -> Optional[Dict]:
    """
    检查 display_name 是否是该实体的参照字段（通过外键指向的表中查找）

    Returns: 找到时返回参照结果，否则返回 None
    """
    foreign_keys = entity.get('foreignKeys', [])
    for fk in foreign_keys:
        ref_uri = fk.get('refUri', '')
        if not ref_uri:
            continue

        ref_entity = find_entity_by_uri(entities, ref_uri)
        if not ref_entity:
            ref_table = _guess_table_from_uri(ref_uri)
            if ref_table:
                ref_entity = find_entity_by_table_name(entities, ref_table)

        if ref_entity:
            for attr in ref_entity.get('attributes', []):
                if attr.get('displayName') == display_name:
                    return {
                        'original': display_name,
                        'displayName': display_name,
                        'found': True,
                        'dbColumnName': attr.get('dbColumnName'),
                        'tableName': attr.get('tableName') or ref_entity.get('tableName'),
                        'entityUri': ref_entity.get('uri'),
                        'viaReference': True,
                        'referenceInfo': {
                            'sourceEntity': entity.get('tableName'),
                            'sourceFK': fk.get('columnName'),
                            'refEntity': ref_entity.get('tableName'),
                            'refUri': ref_uri,
                        },
                        'error': None,
                    }
    return None


def _find_fk_by_name(
    entity: Dict,
    ref_name: str,
    entities: List[Dict]
) -> Optional[Dict]:
    """根据参照名查找外键信息（用于多级参照路径）"""
    fks = entity.get('foreignKeys', [])
    ref_name_lower = ref_name.lower()

    for fk in fks:
        col = fk.get('columnName', '').lower()
        ref_uri = fk.get('refUri', '').lower()

        if ref_name_lower in col or col in ref_name_lower:
            ref_uri_val = fk.get('refUri', '')
            ref_entity = find_entity_by_uri(entities, ref_uri_val)
            if not ref_entity:
                ref_table = _guess_table_from_uri(ref_uri_val)
                if ref_table:
                    ref_entity = find_entity_by_table_name(entities, ref_table)
            if ref_entity:
                return {'fk': fk, 'refEntity': ref_entity}

        if ref_name_lower in ref_uri:
            ref_uri_val = fk.get('refUri', '')
            ref_entity = find_entity_by_uri(entities, ref_uri_val)
            if not ref_entity:
                ref_table = _guess_table_from_uri(ref_uri_val)
                if ref_table:
                    ref_entity = find_entity_by_table_name(entities, ref_table)
            if ref_entity:
                return {'fk': fk, 'refEntity': ref_entity}

    return None


def _find_reference_path(
    field_path: str,
    parts: List[str],
    entities: List[Dict]
) -> Optional[Dict]:
    """
    解析多级参照路径，如 "收款单.客户档案.客户编码"

    - parts[0] = "收款单"       -> 找到起始实体
    - parts[1] = "客户档案"     -> 查找该表的外键
    - parts[2] = "客户编码"      -> 在外键指向的表中查找
    """
    if len(parts) < 2:
        return None

    # 找起始实体（第一段）
    start_bill = parts[0]
    start_entity = None
    for entity in entities:
        bill_name = entity.get('billName', '')
        table_name = entity.get('tableName', '')
        if (start_bill in bill_name or start_bill in table_name or
                bill_name in start_bill or table_name in start_bill):
            start_entity = entity
            break

    if not start_entity:
        return None

    target_field = parts[-1]

    # 两段：业务对象.字段
    if len(parts) == 2:
        for attr in start_entity.get('attributes', []):
            if attr.get('displayName') == target_field:
                return {
                    'original': field_path,
                    'displayName': target_field,
                    'found': True,
                    'dbColumnName': attr.get('dbColumnName'),
                    'tableName': start_entity.get('tableName'),
                    'entityUri': start_entity.get('uri'),
                    'viaReference': False,
                    'referenceInfo': None,
                    'error': None,
                }
        ref_result = _check_reference_field(start_entity, target_field, entities)
        if ref_result:
            ref_result['original'] = field_path
            return ref_result
        return None

    # 三段或更多：业务对象.参照.字段（追踪外键链）
    current_entity = start_entity
    for i in range(1, len(parts) - 1):
        ref_field_name = parts[i]
        fk_match = _find_fk_by_name(current_entity, ref_field_name, entities)
        if not fk_match:
            return None
        current_entity = fk_match['refEntity']

    # 在最终实体中查找目标字段
    for attr in current_entity.get('attributes', []):
        if attr.get('displayName') == target_field:
            return {
                'original': field_path,
                'displayName': target_field,
                'found': True,
                'dbColumnName': attr.get('dbColumnName'),
                'tableName': current_entity.get('tableName'),
                'entityUri': current_entity.get('uri'),
                'viaReference': True,
                'referenceInfo': {
                    'sourceEntity': start_entity.get('tableName'),
                    'refEntity': current_entity.get('tableName'),
                },
                'error': None,
            }
    return None


def verify_field_path(
    field_path: str,
    entities: List[Dict[str, Any]],
    dn_index: Dict[str, List[Dict]] = None,
) -> Dict[str, Any]:
    """
    验证字段路径，支持多种格式：

    1. 单字段: "收款日期"                         -> 在所有实体中查找 displayName
    2. 两段路径: "收款单.单据日期"                 -> 在指定实体中查找
    3. 三段路径: "收款单.客户档案.客户编码"        -> 解析外键链
    4. 直接字段: "bill_date"                       -> 直接作为 dbColumnName 查找

    Returns:
        {
            'original': '收款日期',
            'displayName': '收款日期',
            'found': True/False,
            'dbColumnName': 'bill_date',
            'tableName': 'stwb_settleapply',
            'entityUri': '...',
            'viaReference': True/False,
            'referenceInfo': {...},
            'error': None/'字段不存在'
        }
    """
    result: Dict[str, Any] = {
        'original': field_path,
        'displayName': field_path,
        'found': False,
        'dbColumnName': None,
        'tableName': None,
        'entityUri': None,
        'viaReference': False,
        'referenceInfo': None,
        'error': None,
    }

    parts = field_path.split('.')
    display_name = parts[-1]

    # 多级参照路径（3段及以上）
    if len(parts) >= 3:
        path_result = _find_reference_path(field_path, parts, entities)
        if path_result:
            return path_result

    # 两段路径：业务对象.字段
    if len(parts) == 2:
        target_bill = parts[0]
        for entity in entities:
            bill_name = entity.get('billName', '')
            table_name = entity.get('tableName', '')
            if (target_bill in bill_name or target_bill in table_name or
                    bill_name in target_bill or table_name in target_bill):
                # 先在当前表直接查找
                for attr in entity.get('attributes', []):
                    if attr.get('displayName') == display_name:
                        result.update({
                            'displayName': display_name,
                            'found': True,
                            'dbColumnName': attr.get('dbColumnName'),
                            'tableName': table_name,
                            'entityUri': entity.get('uri'),
                        })
                        return result
                # 尝试参照字段
                ref_result = _check_reference_field(entity, display_name, entities)
                if ref_result:
                    return ref_result

    # 单字段或两段未命中：使用 displayName 索引精确查找
    if dn_index and display_name in dn_index:
        matches = dn_index[display_name]
        if len(matches) == 1:
            m = matches[0]
            result.update({
                'displayName': display_name,
                'found': True,
                'dbColumnName': m['dbColumnName'],
                'tableName': m['tableName'],
                'entityUri': m.get('entityUri'),
            })
        else:
            # 多个匹配，返回第一个（带 warning）
            m = matches[0]
            result.update({
                'displayName': display_name,
                'found': True,
                'dbColumnName': m['dbColumnName'],
                'tableName': m['tableName'],
                'entityUri': m.get('entityUri'),
                'error': f"字段 '{display_name}' 在元数据中有多处定义，返回第一个匹配（共 {len(matches)} 处）",
            })
        return result

    # 最后兜底：遍历所有实体
    for entity in entities:
        table_name = entity.get('tableName', '')
        for attr in entity.get('attributes', []):
            if attr.get('displayName') == display_name:
                result.update({
                    'displayName': display_name,
                    'found': True,
                    'dbColumnName': attr.get('dbColumnName'),
                    'tableName': table_name,
                    'entityUri': entity.get('uri'),
                })
                return result

    result['error'] = f"字段 '{display_name}' 在元数据中不存在"
    return result


def verify_fields(
    fields: List[str],
    entities: List[Dict[str, Any]],
    dn_index: Dict[str, List[Dict]] = None,
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """批量验证字段"""
    verified, unverified = [], []
    for field in fields:
        result = verify_field_path(field, entities, dn_index)
        if result['found']:
            verified.append(result)
        else:
            unverified.append(result)
    return verified, unverified


# =============================================================================
# 第三部分：字段校验（验证能力）
# =============================================================================

# 平台隐含字段：这些字段是隐式筛选条件，不需要出现在 SELECT 中
_IMPLICIT_WHERE_FIELDS: Set[str] = {
    'dr', 'ytenant_id', 'tenant_id', 'pubts',
    'create_time', 'creator', 'modifier', 'modify_time',
    'auditor', 'audit_time', 'ts',
}


def _is_implicit_field(col: str) -> bool:
    """判断是否为平台隐含字段"""
    return col.lower() in _IMPLICIT_WHERE_FIELDS


def _levenshtein_distance(a: str, b: str) -> int:
    """计算编辑距离（Levenshtein）"""
    if len(a) > len(b):
        a, b = b, a
    distances = range(len(a) + 1)
    for i2, c2 in enumerate(b):
        distances_ = [i2 + 1]
        for i1, c1 in enumerate(a):
            distances_.append(
                distances[i1] if c1 == c2 else 1 + min(distances[i1], distances[i1 + 1], distances_[-1])
            )
        distances = distances_
    return distances[-1]


def _find_similar_columns(target: str, candidates: Set[str]) -> List[str]:
    """
    模糊匹配：查找与目标字段相似的候选字段

    相似规则：前缀差异、包含关系、Levenshtein 距离 <= 2
    """
    target_lower = target.lower()
    suggestions: List[Tuple[str, str]] = []

    for col in candidates:
        col_lower = col.lower()
        target_stripped = target_lower
        col_stripped = col_lower

        # 去掉常见前缀后比较
        for p in ('i', 'c', 'v', 'n', 'b'):
            if target_lower.startswith(p) and len(target_lower) > len(p):
                target_stripped = target_lower[len(p):]
            if col_lower.startswith(p) and len(col_lower) > len(p):
                col_stripped = col_lower[len(p):]

        if target_stripped == col_stripped:
            suggestions.append((col, 'prefix_diff'))
        elif target_lower in col_lower or col_lower in target_lower:
            suggestions.append((col, 'substring'))
        elif _levenshtein_distance(target_lower, col_lower) <= 2:
            suggestions.append((col, 'similar'))

    priority = {'prefix_diff': 0, 'similar': 1, 'substring': 2}
    suggestions.sort(key=lambda x: priority.get(x[1], 3))
    return [s[0] for s in suggestions]


def validate_field_against_metadata(
    table_name: str,
    db_column: str,
    entities_index: Dict[str, Set[str]],
) -> Optional[Dict]:
    """
    校验 tableName.dbColumnName 是否在元数据中存在

    Returns: None = 通过，Dict = 失败信息
    """
    if not table_name or not db_column or table_name == '-' or db_column == '-':
        return None  # 跳过聚合字段等占位符

    if 'COUNT' in db_column.upper() or '聚合' in db_column:
        return None

    # 找到匹配的表
    matching_keys = [k for k in entities_index.keys() if k.endswith(f".{table_name}")]
    if not matching_keys:
        return {
            'table': table_name,
            'column': db_column,
            'error': f"表 '{table_name}' 在元数据中不存在",
            'suggestion': None,
            'actual_columns': [],
        }

    actual_columns = entities_index[matching_keys[0]]
    if db_column in actual_columns:
        return None

    suggestions = _find_similar_columns(db_column, actual_columns)
    return {
        'table': table_name,
        'column': db_column,
        'error': f"字段 '{db_column}' 在表 '{table_name}' 中不存在",
        'suggestion': suggestions[0] if suggestions else None,
        'actual_columns': list(actual_columns)[:15],
    }


def validate_mapping_table(
    entities_path: str,
    mapping_path: str,
) -> Dict:
    """
    校验字段映射表（md）中每个 tableName.dbColumnName 是否真实存在
    """
    entities = load_entities(entities_path)
    entities_index = build_column_index(entities)

    fields = parse_mapping_table(mapping_path)
    errors = []
    matched = 0

    for field in fields:
        table = field.get('table_name', '')
        db_col = field.get('db_column', '')
        if not table or not db_col:
            continue
        result = validate_field_against_metadata(table, db_col, entities_index)
        if result is None:
            matched += 1
        else:
            errors.append(result)

    return {
        'valid': len(errors) == 0,
        'total': len(fields),
        'matched': matched,
        'errors': errors,
    }


# =============================================================================
# 第四部分：映射表解析
# =============================================================================

def extract_select_fields(mapping_content: str) -> List[str]:
    """从字段映射表的「字段映射」区域提取 SELECT 字段列表"""
    select_fields = []
    in_field_section = False

    for line in mapping_content.split('\n'):
        if '报表字段' in line and '|' in line:
            in_field_section = True
            continue
        if in_field_section and ('---' in line or '|--' in line):
            continue
        if in_field_section and line.startswith('## '):
            break
        if in_field_section and '|' in line:
            parts = [p.strip() for p in line.split('|')]
            if len(parts) >= 3:
                for p in parts:
                    p = p.strip()
                    if not p or p in ('报表字段', '数据库表.字段', '字段类型', '说明', '---'):
                        continue
                    select_fields.append(p)
                    break
    return select_fields


def extract_where_fields(mapping_content: str) -> List[Dict]:
    """从字段映射表的「筛选条件」区域提取字段引用"""
    where_fields = []
    in_where_section = False

    for line in mapping_content.split('\n'):
        if '筛选条件' in line or 'WHERE' in line.upper():
            in_where_section = True
            continue
        if in_where_section and line.startswith('## '):
            in_where_section = False
            break
        if in_where_section and line.strip():
            field_pattern = re.compile(r'([a-zA-Z_]\w*)\.([a-zA-Z_]\w*)')
            for m in field_pattern.finditer(line):
                alias, col = m.group(1), m.group(2)
                if col.upper() in ('AND', 'OR', 'NOT', 'IN', 'LIKE', 'BETWEEN', 'IS', 'NULL'):
                    continue
                where_fields.append({'field': f"{alias}.{col}", 'from_where_section': True})
            standalone_pattern = re.compile(
                r'^[\s\-]*([a-zA-Z_]\w+)\s+(?:过滤|=|!=|<|>|<=|>=|IN|LIKE)', re.MULTILINE
            )
            for m in standalone_pattern.finditer(line):
                col = m.group(1)
                if col.upper() not in ('AND', 'OR', 'NOT', 'IN', 'LIKE', 'BETWEEN', 'IS', 'NULL', '过滤'):
                    where_fields.append({'field': col, 'from_where_section': True})

    return where_fields


def extract_table_aliases_from_mapping(mapping_content: str) -> Dict[str, str]:
    """从 JOIN/FROM 行提取别名 -> 表名映射"""
    aliases = {}
    join_pattern = re.compile(r'(?:FROM|JOIN)\s+(\w+)\s+(?:AS\s+)?(\w+)', re.IGNORECASE)
    for line in mapping_content.split('\n'):
        m = join_pattern.search(line)
        if m:
            table, alias = m.group(1), m.group(2)
            aliases[alias] = table
    return aliases


def parse_mapping_table(mapping_path: str) -> List[Dict]:
    """
    解析字段映射表 Markdown

    支持格式:
    | 序号 | 报表字段 | tableName | dbColumnName | 字段类型 | 说明 |
    """
    with open(mapping_path, 'r', encoding='utf-8') as f:
        content = f.read()

    fields = []
    in_field_section = False

    for line in content.split('\n'):
        if 'tableName' in line and 'dbColumnName' in line:
            in_field_section = True
            continue
        if '---' in line or '|--' in line:
            continue
        if in_field_section and '|' in line:
            parts = [p.strip() for p in line.split('|')]
            # 如果第二个非空元素不是数字，说明到了其他区域
            non_empty = [p for p in parts if p and p not in ('---', '|--')]
            if len(non_empty) >= 2 and not non_empty[0].isdigit():
                continue
            if len(parts) >= 5 and parts[1].isdigit():
                fields.append({
                    'seq': parts[1],
                    'report_field': parts[2] if len(parts) > 2 else '',
                    'table_name': parts[3] if len(parts) > 3 else '',
                    'db_column': parts[4] if len(parts) > 4 else '',
                    'field_type': parts[5] if len(parts) > 5 else '',
                })
    return fields


# =============================================================================
# 第五部分：WHERE 字段校验
# =============================================================================

def validate_where_fields_in_select(mapping_path: str) -> Dict:
    """
    校验 WHERE 条件字段是否出现在 SELECT 中

    规则：
      - 平台隐含字段（dr, ytenant_id 等）→ 跳过
      - 来自「筛选条件」区域的字段 → 警告但不阻塞
      - 其他 WHERE 字段 → 必须出现在 SELECT 中
    """
    with open(mapping_path, 'r', encoding='utf-8') as f:
        content = f.read()

    where_fields_data = extract_where_fields(content)
    select_fields = extract_select_fields(content)

    missing = []
    query_condition_warnings = []

    for item in where_fields_data:
        wf = item['field']
        from_where_section = item.get('from_where_section', False)

        if '.' in wf:
            alias, col = wf.split('.', 1)
        else:
            alias, col = '', wf

        if _is_implicit_field(col):
            continue

        found = any(col in sf or sf in col for sf in select_fields)

        if not found:
            if from_where_section:
                query_condition_warnings.append({
                    'field': wf,
                    'column': col,
                    'reason': f"查询条件字段 '{wf}' 不在 SELECT 报表字段中",
                    'suggestion': "可忽略；如需显示请添加到 SELECT",
                })
            else:
                missing.append({
                    'field': wf,
                    'column': col,
                    'reason': f"WHERE 条件字段 '{wf}' 未出现在 SELECT 报表字段中",
                    'suggestion': "请添加到 SELECT 或修改 WHERE 条件",
                })

    return {
        'valid': len(missing) == 0,
        'total_where_fields': len(where_fields_data),
        'select_fields': select_fields,
        'missing_fields': missing,
        'query_condition_warnings': query_condition_warnings,
    }


# =============================================================================
# 第六部分：Excel 支持
# =============================================================================

def parse_excel_fields(excel_path: str) -> List[str]:
    """从 Excel 提取数据源字段"""
    try:
        import openpyxl
    except ImportError:
        print("错误: 需要 openpyxl 库来读取 Excel 文件", file=sys.stderr)
        print("  pip install openpyxl", file=sys.stderr)
        return []

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    fields = []

    for sheet in wb.worksheets:
        # 跳过说明sheet
        if sheet.title in ('说明', 'Sheet1') and sheet.max_row < 5:
            continue
        for row in sheet.iter_rows(values_only=True):
            if not row or not row[0]:
                continue
            cell = str(row[0]).strip()
            # 排除序号、标题行等
            if cell.isdigit() and len(str(row[1] or '').strip()) > 0:
                # 可能是序号行，取第二列作为字段
                field = str(row[1]).strip() if len(row) > 1 else ''
                if field and field not in fields:
                    fields.append(field)
            elif cell and cell not in fields:
                # 直接是字段名
                if len(cell) > 1 and not cell.startswith('#'):
                    fields.append(cell)

    wb.close()
    return fields


# =============================================================================
# 第七部分：报告输出
# =============================================================================

def print_field_discovery_report(verified: List, unverified: List):
    """打印字段发现报告（来自 Excel 或字段列表）"""
    print("\n" + "=" * 70)
    print("字段发现报告")
    print("=" * 70)

    print(f"\n[OK] 已找到: {len(verified)} 个")
    print("-" * 70)
    for v in verified:
        ref_info = ""
        if v.get('viaReference'):
            ri = v.get('referenceInfo', {})
            ref_info = f" [via JOIN: {ri.get('sourceEntity')}.{ri.get('sourceFK')} -> {ri.get('refEntity')}]"
        warn = " ⚠️" if v.get('error') and '多处定义' in v.get('error', '') else ""
        print(f"  {v['displayName']:<20} -> {v['dbColumnName']:<28} (table: {v['tableName']}){ref_info}{warn}")

    if unverified:
        print(f"\n[FAIL] 未找到: {len(unverified)} 个")
        print("-" * 70)
        for u in unverified:
            print(f"  {u['original']:<30}  [X] {u['error']}")
        print("\n建议:")
        print("  1. 检查字段显示名是否正确")
        print("  2. 确认该字段是否属于已拉取的业务对象")
        print("  3. 可能需要补充拉取相关业务对象的元数据")
    else:
        print(f"\n[PASS] 所有 {len(verified)} 个字段验证通过!")
    print("=" * 70)


def print_mapping_validation_result(result: Dict):
    """打印映射表校验结果"""
    print(f"\n{'=' * 60}")
    print(f"[PASS] 字段映射表校验结果")
    print(f"{'=' * 60}")
    print(f"总计字段: {result['total']}")
    print(f"已匹配:   {result['matched']} [OK]")
    print(f"未匹配:   {len(result['errors'])} [FAIL]")

    if result['valid']:
        print(f"\n[PASS] 所有字段校验通过!")
        return 0

    print(f"\n[FAIL] 发现 {len(result['errors'])} 个未匹配的字段:\n")
    for i, err in enumerate(result['errors'], 1):
        print(f"{i}. [FAIL] 表 '{err['table']}' 的字段 '{err['column']}' 不存在")
        if err.get('suggestion'):
            print(f"   [SUGGEST] 建议修正为: '{err['suggestion']}'")
        if err.get('actual_columns'):
            print(f"   [ACTUAL] 表中实际字段: {', '.join(err['actual_columns'][:10])}")
        print()
    return 1


def print_where_validation_result(where_result: Dict) -> int:
    """打印 WHERE 字段校验结果"""
    print(f"\n{'=' * 60}")
    print(f"[CHECK] WHERE 字段 → SELECT 校验")
    print(f"{'=' * 60}")
    print(f"WHERE 字段总数: {where_result['total_where_fields']}")
    sel = where_result['select_fields']
    print(f"SELECT 报表字段: {', '.join(sel[:10])}"
          f"{'...' if len(sel) > 10 else ''}")

    exit_code = 0
    if where_result['valid']:
        print(f"\n[PASS] 所有 WHERE 字段均已在 SELECT 中!")
    else:
        print(f"\n[FAIL] 发现 {len(where_result['missing_fields'])} 个必须修复的字段:")
        for i, m in enumerate(where_result['missing_fields'], 1):
            print(f"  {i}. {m['field']}")
            print(f"     原因: {m['reason']}")
            print(f"     建议: {m['suggestion']}")
        exit_code = 1

    if where_result.get('query_condition_warnings'):
        print(f"\n[WARN] 发现 {len(where_result['query_condition_warnings'])} 个查询条件字段不在 SELECT 中:")
        for i, w in enumerate(where_result['query_condition_warnings'], 1):
            print(f"  {i}. {w['field']} - 查询条件字段，可忽略")

    print("=" * 60)
    return exit_code


# =============================================================================
# 主入口
# =============================================================================

def main() -> int:
    parser = argparse.ArgumentParser(
        description='字段预检验脚本 v2.0（合并 phase1 + phase2 能力）',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
    # 从 Excel 提取字段并验证元数据
    python field_verify.py --entities output/entities.json --excel 报表模板.xlsx

    # 验证字段列表
    python field_verify.py --entities output/entities.json --fields 收款日期 收款金额

    # 校验映射表中的字段是否在元数据中存在
    python field_verify.py --entities output/entities.json --mapping output/字段映射表.md

    # 校验 WHERE 字段是否出现在 SELECT 中
    python field_verify.py --mapping output/字段映射表.md --check-where

    # 完整校验（映射表 + WHERE）
    python field_verify.py --entities output/entities.json --mapping output/字段映射表.md --check-where
        """
    )
    parser.add_argument('--entities', help='entities.json 路径')
    parser.add_argument('--mapping', help='字段映射表.md 路径')
    parser.add_argument('--fields', nargs='+', help='要验证的字段列表')
    parser.add_argument('--excel', help='Excel 文件路径（自动提取字段）')
    parser.add_argument('--check-where', action='store_true',
                        help='校验 WHERE 条件字段是否出现在 SELECT 中')
    parser.add_argument('--output-report', help='输出 JSON 报告到文件')

    args = parser.parse_args()

    # 参数检查
    has_entities = bool(args.entities)
    has_mapping = bool(args.mapping)
    has_fields = bool(args.fields)
    has_excel = bool(args.excel)

    if not has_entities and not has_mapping and not has_fields and not has_excel:
        print("错误: 请指定 --entities / --mapping / --fields / --excel 之一", file=sys.stderr)
        return 1

    exit_code = 0
    all_results = {}

    # 预加载 entities
    entities: List[Dict] = []
    entities_index: Dict[str, Set[str]] = {}
    dn_index: Dict[str, List[Dict]] = {}
    if has_entities:
        entities_path = Path(args.entities)
        if not entities_path.exists():
            print(f"错误: entities.json 不存在: {args.entities}", file=sys.stderr)
            return 1
        entities = load_entities(args.entities)
        entities_index = build_column_index(entities)
        dn_index = build_displayname_index(entities)
        print(f"已加载 {len(entities)} 个实体")

    # 1. WHERE 字段校验
    if args.check_where and has_mapping:
        mapping_path = Path(args.mapping)
        if not mapping_path.exists():
            print(f"错误: 映射表不存在: {args.mapping}", file=sys.stderr)
            return 1
        where_result = validate_where_fields_in_select(args.mapping)
        all_results['where_validation'] = where_result
        exit_code = max(exit_code, print_where_validation_result(where_result))

    # 2. 映射表字段校验
    if has_entities and has_mapping:
        mapping_path = Path(args.mapping)
        if not mapping_path.exists():
            print(f"错误: 映射表不存在: {args.mapping}", file=sys.stderr)
            return 1
        mapping_result = validate_mapping_table(args.entities, args.mapping)
        all_results['mapping_validation'] = mapping_result
        exit_code = max(exit_code, print_mapping_validation_result(mapping_result))

    # 3. 字段发现（从 Excel 或字段列表）
    if (has_fields or has_excel) and has_entities:
        fields_to_verify = []
        if has_fields:
            fields_to_verify = args.fields
        elif has_excel:
            excel_path = Path(args.excel)
            if not excel_path.exists():
                print(f"错误: Excel 文件不存在: {args.excel}", file=sys.stderr)
                return 1
            fields_to_verify = parse_excel_fields(args.excel)
            print(f"从 Excel 提取 {len(fields_to_verify)} 个字段")

        verified, unverified = verify_fields(fields_to_verify, entities, dn_index)
        all_results['field_discovery'] = {'verified': verified, 'unverified': unverified}
        print_field_discovery_report(verified, unverified)
        if unverified:
            exit_code = 1

    # 输出报告
    if args.output_report and all_results:
        with open(args.output_report, 'w', encoding='utf-8') as f:
            json.dump(all_results, f, ensure_ascii=False, indent=2)
        print(f"\n报告已保存到: {args.output_report}")

    return exit_code


if __name__ == '__main__':
    sys.exit(main())
